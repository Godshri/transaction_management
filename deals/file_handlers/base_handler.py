import abc
import csv
import io
from django.http import HttpResponse
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


class BaseFileHandler(abc.ABC):
    """Базовый класс для обработки файлов"""

    @abc.abstractmethod
    def read_records(self, file):
        """Чтение записей из файла"""
        pass

    @abc.abstractmethod
    def write_records(self, records):
        """Запись записей в файл"""
        pass


class CSVHandler(BaseFileHandler):
    """Обработчик CSV файлов"""

    def read_records(self, file):
        """Чтение CSV файла с поддержкой русской кодировки"""
        records = []

        try:

            for encoding in ['utf-8-sig', 'utf-8', 'cp1251', 'windows-1251']:
                try:
                    file.seek(0)
                    content = file.read().decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise ValueError("Не удалось определить кодировку файла")


            sample = content[:1024]
            if ';' in sample and sample.count(';') > sample.count(','):
                delimiter = ';'
            else:
                delimiter = ','

            io_string = io.StringIO(content)


            reader = csv.DictReader(io_string, delimiter=delimiter)

            for row_num, row in enumerate(reader, 1):
                try:

                    normalized_row = {}
                    for key, value in row.items():
                        if key:
                            normalized_key = key.strip().lower().replace('\ufeff', '')
                            normalized_row[normalized_key] = value.strip() if value else ''


                    record = {
                        'first_name': '',
                        'last_name': '',
                        'phone': '',
                        'email': '',
                        'company_name': ''
                    }


                    for key in normalized_row.keys():
                        if any(x in key for x in ['имя', 'name', 'first']):
                            record['first_name'] = normalized_row[key]
                        elif any(x in key for x in ['фамилия', 'last', 'surname']):
                            record['last_name'] = normalized_row[key]
                        elif any(x in key for x in ['телефон', 'phone', 'тел']):
                            record['phone'] = normalized_row[key]
                        elif any(x in key for x in ['почта', 'email', 'mail']):
                            record['email'] = normalized_row[key]
                        elif any(x in key for x in ['компания', 'company']):
                            record['company_name'] = normalized_row[key]


                    if record['first_name'] or record['last_name']:
                        records.append(record)

                except Exception as e:
                    print(f"Ошибка в строке {row_num}: {e}")
                    continue

        except Exception as e:
            print(f"Ошибка чтения CSV: {e}")
            raise

        return records

    def write_records(self, records):
        """Создание CSV файла с русскими заголовками"""
        try:
            output = io.StringIO()
            writer = csv.writer(output, delimiter=',', quoting=csv.QUOTE_ALL)


            writer.writerow(['Имя', 'Фамилия', 'Телефон', 'Email', 'Компания'])


            for record in records:
                writer.writerow([
                    record.get('first_name', ''),
                    record.get('last_name', ''),
                    record.get('phone', ''),
                    record.get('email', ''),
                    record.get('company_name', '')
                ])

            response = HttpResponse(
                output.getvalue().encode('utf-8-sig'),
                content_type='text/csv; charset=utf-8-sig'
            )
            response['Content-Disposition'] = 'attachment; filename="contacts_export.csv"'

            return response

        except Exception as e:
            print(f"Ошибка записи CSV: {e}")
            raise


class XLSXHandler(BaseFileHandler):
    """Обработчик XLSX файлов"""

    def read_records(self, file):
        """Чтение XLSX файла"""
        records = []

        try:
            file.seek(0)
            workbook = openpyxl.load_workbook(file, read_only=True)
            sheet = workbook.active

            headers = []

            for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row_num == 1:

                    headers = [str(cell).strip().lower() if cell else '' for cell in row]
                    continue

                if not any(cell for cell in row):
                    continue

                try:
                    record = {
                        'first_name': '',
                        'last_name': '',
                        'phone': '',
                        'email': '',
                        'company_name': ''
                    }


                    for col_num, cell in enumerate(row):
                        if col_num < len(headers) and cell is not None:
                            cell_value = str(cell).strip()
                            header = headers[col_num]

                            if any(x in header for x in ['имя', 'name', 'first']):
                                record['first_name'] = cell_value
                            elif any(x in header for x in ['фамилия', 'last', 'surname']):
                                record['last_name'] = cell_value
                            elif any(x in header for x in ['телефон', 'phone', 'тел']):
                                record['phone'] = cell_value
                            elif any(x in header for x in ['почта', 'email', 'mail']):
                                record['email'] = cell_value
                            elif any(x in header for x in ['компания', 'company']):
                                record['company_name'] = cell_value


                    if record['first_name'] or record['last_name']:
                        records.append(record)

                except Exception as e:
                    print(f"Ошибка в строке {row_num}: {e}")
                    continue

            workbook.close()

        except InvalidFileException:
            raise ValueError("Некорректный XLSX файл")
        except Exception as e:
            print(f"Ошибка чтения XLSX: {e}")
            raise

        return records

    def write_records(self, records):
        """Создание XLSX файла"""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Контакты"


            sheet.append(['Имя', 'Фамилия', 'Телефон', 'Email', 'Компания'])


            for record in records:
                sheet.append([
                    record.get('first_name', ''),
                    record.get('last_name', ''),
                    record.get('phone', ''),
                    record.get('email', ''),
                    record.get('company_name', '')
                ])


            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="contacts_export.xlsx"'

            workbook.save(response)
            return response

        except Exception as e:
            print(f"Ошибка записи XLSX: {e}")
            raise


class FileHandlerFactory:
    """Фабрика для получения обработчиков файлов"""

    @staticmethod
    def get_handler(file_format):
        file_format = file_format.lower()
        if file_format == 'csv':
            return CSVHandler()
        elif file_format == 'xlsx':
            return XLSXHandler()
        else:
            raise ValueError(f"Неподдерживаемый формат файла: {file_format}")